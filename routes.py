from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload, selectinload
from sqlalchemy import and_, or_, func
from datetime import datetime, timedelta
from typing import Optional, List
from pydantic import BaseModel
from database import DEFAULT_EMPLOYEE_PASSWORD, get_db, SessionLocal, ensure_employee_credentials
from models import *
from auth import AUTH_MODE, get_current_user, require_role, verify_password, create_access_token, hash_password, validate_microsoft_id_token
from email_utils import send_notification_email, send_notification_email_batch, send_password_reset_email, mark_email_sent
from nest_employee_sync import sync_nest_employees
import hashlib
import os
import secrets
import uuid
import requests
import re
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO

router = APIRouter()
email_executor = ThreadPoolExecutor(max_workers=4)
REMEMBER_ME_TOKEN_EXPIRE_DAYS = int(os.getenv("REMEMBER_ME_TOKEN_EXPIRE_DAYS", "30"))
EMAIL_ENABLED_EVENT_TYPES = {"step1_activated", "step1_deadline_alert", "kra_sent_back"}

PUBLIC_TO_DB_CYCLE_PERIOD = {
    "H1_SEP": "H1_MAR",
    "H2_MAR": "H2_SEP",
    "H1_MAR": "H1_MAR",
    "H2_SEP": "H2_SEP",
}

DB_TO_PUBLIC_CYCLE_PERIOD = {
    "H1_MAR": "H1_SEP",
    "H2_SEP": "H2_MAR",
    "H1_SEP": "H1_SEP",
    "H2_MAR": "H2_MAR",
}


def _login_token_for(emp: Employee, remember_me: bool = False):
    expires_delta = timedelta(days=REMEMBER_ME_TOKEN_EXPIRE_DAYS) if remember_me else None
    return create_access_token(
        {"sub": str(emp.id), "role": emp.role.value, "name": emp.full_name},
        expires_delta=expires_delta,
    )


def _format_period(p):
    raw = p.value if hasattr(p, 'value') else p
    if raw in {"H2_MAR", "H2_SEP"}:
        return "Mar (H2)"
    if raw in {"H1_SEP", "H1_MAR"}:
        return "Sep (H1)"
    return str(raw or "")

def _cycle_label(cycle_name: str, period: str, bold: bool = False) -> str:
    label = f"{cycle_name} {_format_period(period)}"
    return f"**{label}**" if bold else label

def _cycle_close_alert_body(cycle_name: str, period: str, deadline: str) -> str:
    cycle_label = _cycle_label(cycle_name, period, bold=True)
    return (
        f"This is a reminder that the performance cycle {cycle_label} is scheduled for closure on {deadline} in the EDGE portal. "
        f"Please review and complete any pending actions related to KRA submission, approvals, self-rating, or final rating before this deadline.\n\n"
        f"If you have already completed your pending activities in EDGE, please ignore this reminder. "
        f"Once the cycle is closed by HR/Admin, further updates will not be allowed unless the review window is reopened by the administrator.\n\n"
    )

def _cycle_closed_body(cycle_name: str, period: str, is_employee: bool = False) -> str:
    cycle_label = _cycle_label(cycle_name, period, bold=True)
    first_line = (
        f"The performance cycle {cycle_label} has now been closed in EDGE. "
        f"Your final performance outcome is available for review."
        if is_employee
        else f"The performance cycle {cycle_label} has now been closed in EDGE. Final performance outcomes have been released."
    )
    return (
        f"{first_line}\n\n"
        f"If you have already reviewed the final status and related performance details in EDGE, please ignore this message. "
        f"If not, please log in to the application and review the cycle summary at your earliest convenience.\n\n"
    )

def _step1_activated_body(cycle_name: str, period: str, deadline: str) -> str:
    cycle_label = _cycle_label(cycle_name, period, bold=True)
    return f"The performance cycle '{cycle_label}' has been activated for KRAs Allocation.\n\nManagers and Approvers are requested to complete KRAs allocation and baselining for their team members in the EDGE portal. The deadline for KRAs submission is {deadline}."

def _step1_deadline_alert_body(cycle_name: str, period: str, deadline: str) -> str:
    cycle_label = _cycle_label(cycle_name, period, bold=True)
    return f"Reminder: The deadline for KRA Allocation in performance cycle {cycle_label} is approaching on {deadline}.\n\nManagers and Approvers who have not yet finalized KRA allocations for their reportees are requested to do so at the earliest to avoid delays in the performance cycle."

def _step2_activated_body(cycle_name: str, period: str, self_deadline: str, mgr_deadline: str) -> str:
    cycle_label = _cycle_label(cycle_name, period, bold=True)
    return f"Performance Rating is now open for the performance cycle {cycle_label}.\n\n- Employees: Please complete your self-rating by {self_deadline}.\n- Managers: Please complete manager ratings by {mgr_deadline}."

def _step2_deadline_alert_body(cycle_name: str, period: str, self_deadline: str, mgr_deadline: str) -> str:
    cycle_label = _cycle_label(cycle_name, period, bold=True)
    return f"Reminder: The deadlines for Performance Rating in performance cycle {cycle_label} are approaching.\n\n- Self-rating deadline: {self_deadline}\n- Manager rating deadline: {mgr_deadline}\n\nIf you have pending ratings or comments, please complete them in the EDGE portal immediately."

def _rating_finalised_body(cycle_name: str, period: str, employee_name: str = "", manager_name: str | None = None, is_employee: bool = False) -> str:
    cycle_label = _cycle_label(cycle_name, period, bold=True)
    if is_employee and manager_name:
        message = (
            f"Your performance rating for cycle {cycle_label} shared by {manager_name} has been finalised successfully."
        )
    elif is_employee:
        message = (
            f"Your performance rating for cycle {cycle_label} has been finalised successfully."
        )
    else:
        message = (
            f"The performance rating for cycle {cycle_label} has been reviewed and finalised by the approver."
        )

    return f"""
{message}

You may now log in to the PMS portal to review the final status and feedback.

Please note that once the cycle is closed by HR/Admin, further updates will not be allowed unless the review window is reopened by the administrator.
""".strip()


def _clean_text(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    value = value.strip()
    return value or None

def _normalize_excel_header(value) -> str:
    return "".join(ch for ch in str(value or "").strip().lower() if ch.isalnum())

def _normalize_band_code(value) -> str:
    return "".join(ch for ch in str(value or "").strip().lower() if ch.isalnum())

def _cell_text(value) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text or None

def _get_excel_value(row: dict, *aliases: str) -> Optional[str]:
    for alias in aliases:
        value = row.get(_normalize_excel_header(alias))
        if value:
            return value
    return None

def _excel_code_seed(value: str) -> str:
    seed = "".join(ch for ch in value.upper() if ch.isalnum())
    return seed[:8] or "KRA"

def _clean_excel_code(value: Optional[str], fallback: str, max_length: int = 20) -> str:
    if value and len(value) <= max_length:
        return value
    return fallback[:max_length]

def _split_key_kpis(value: Optional[str]) -> List[str]:
    if not value:
        return []
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    parts = re.split(r"\n+|;+", text)
    result = []
    for part in parts:
        item = re.sub(r"^\s*(?:[-*•]+|\d+[\.)])\s*", "", part).strip()
        if item:
            result.append(item)
    return result

def _find_excel_header_row(rows) -> tuple[int, List[str]] | None:
    for index, values in enumerate(rows[:10]):
        headers = [_normalize_excel_header(value) for value in values]
        header_set = set(headers)
        if "kra" in header_set and ("keykpis" in header_set or "keykpi" in header_set or "kpi" in header_set or "kpiname" in header_set):
            return index, headers
    return None


def _validate_kra_allocation_payload(db: Session, allocations: List[KRAAllocate], require_complete: bool = True):
    if not allocations:
        raise HTTPException(400, "At least one KRA is required")

    standard_kra_ids = [str(item.kra_master_id) for item in allocations if item.kra_master_id]
    if len(set(standard_kra_ids)) != len(standard_kra_ids):
        raise HTTPException(400, "Duplicate KRAs are not allowed")

    if require_complete:
        total = sum(k.weightage_pct for k in allocations)
        if abs(total - 100.0) > 0.01:
            raise HTTPException(400, f"Weightage must total 100%, got {total}%")

    custom_weight = 0.0
    custom_kra_count = 0

    for item in allocations:
        item.measurement_comment = _clean_text(item.measurement_comment)
        item.custom_kra_name = _clean_text(item.custom_kra_name)
        item.custom_kra_description = _clean_text(item.custom_kra_description)
        item.custom_kpis = [
            CustomKPIAllocate(custom_kpi_name=name)
            for name in [_clean_text(k.custom_kpi_name) for k in item.custom_kpis]
            if name
        ]

        is_custom_kra = bool(item.custom_kra_name)

        if require_complete and item.weightage_pct <= 0:
            raise HTTPException(400, "Weightage must be greater than 0")
        if require_complete and not item.measurement_comment:
            raise HTTPException(400, "Measurement description is mandatory for every KRA")

        if is_custom_kra:
            custom_kra_count += 1
            if custom_kra_count > 1:
                raise HTTPException(400, "Only one custom KRA is allowed per employee")
            custom_weight += item.weightage_pct
            if item.kra_master_id:
                raise HTTPException(400, "Custom KRAs cannot reference a master KRA")
            if item.kpi_ids:
                raise HTTPException(400, "Custom KRAs cannot use KPI master selections")
            if require_complete and not item.custom_kpis:
                raise HTTPException(400, "At least one custom KPI is required for a custom KRA")
            continue

        if item.custom_kpis:
            raise HTTPException(400, "Custom KPI names are allowed only for custom KRAs")
        if not item.kra_master_id:
            if require_complete:
                raise HTTPException(400, "KRA selection is required")
            continue
        if len({str(kpi_id) for kpi_id in item.kpi_ids}) != len(item.kpi_ids):
            raise HTTPException(400, "Duplicate KPIs are not allowed within a KRA")
        if require_complete and not item.kpi_ids:
            raise HTTPException(400, "At least one KPI is required for every KRA")

        kra_master = db.query(KRAMaster).filter(
            KRAMaster.id == item.kra_master_id,
            KRAMaster.is_active == True,
        ).first()
        if not kra_master:
            raise HTTPException(400, "Selected KRA is invalid")

        valid_kpi_ids = {
            str(row.id) for row in db.query(KPIMaster.id).filter(
                KPIMaster.kra_master_id == item.kra_master_id,
                KPIMaster.is_active == True,
            ).all()
        }
        requested_kpi_ids = {str(kpi_id) for kpi_id in item.kpi_ids}
        if not requested_kpi_ids.issubset(valid_kpi_ids):
            raise HTTPException(400, "One or more selected KPIs are invalid for the chosen KRA")

    if custom_weight > 25.0:
        raise HTTPException(400, f"Custom KRA weightage cannot exceed 25%, got {custom_weight}%")

# ─── Schemas ───────────────────────────────────────────────────────────────────

class LoginRequest(BaseModel):
    username: str
    password: str
    remember_me: bool = False


class ChangePasswordRequest(BaseModel):
    confirm_email: str
    current_password: str
    new_password: str


class MicrosoftLoginRequest(BaseModel):
    id_token: str

class ForgotPasswordRequest(BaseModel):
    email: str

class ResetPasswordRequest(BaseModel):
    token: str
    password: str

class CycleCreate(BaseModel):
    cycle_name: str
    financial_year: str
    period: str
    step1_open_date: str
    step1_kra_deadline: str
    step1_approval_date: str
    step2_open_date: Optional[str] = None
    step2_self_deadline: Optional[str] = None
    step2_mgr_deadline: Optional[str] = None
    step2_approval_date: Optional[str] = None

class Step2OpenRequest(BaseModel):
    step2_open_date: str
    step2_self_deadline: str
    step2_mgr_deadline: str
    step2_approval_date: str

class Step1ApprovalDateUpdate(BaseModel):
    step1_approval_date: str

class EmployeeCreate(BaseModel):
    employee_code: str
    full_name: str
    email: str
    band_id: str
    manager_id: Optional[str] = None
    approver_id: Optional[str] = None
    role: str

class CustomKPIAllocate(BaseModel):
    custom_kpi_name: str

class KRAAllocate(BaseModel):
    kra_master_id: Optional[str] = None
    custom_kra_name: Optional[str] = None
    custom_kra_description: Optional[str] = None
    weightage_pct: float
    kpi_ids: List[str] = []
    custom_kpis: List[CustomKPIAllocate] = []
    measurement_comment: str

class KRAAllocationSubmit(BaseModel):
    kras: List[KRAAllocate]

class SelfRatingUpdate(BaseModel):
    kra_id: str
    self_rating: int
    self_comments: str

class ManagerRatingUpdate(BaseModel):
    kra_id: str
    mgr_rating: int
    mgr_comments: str

class ApprovalActionRequest(BaseModel):
    action: str  # approve | send_back
    comment: Optional[str] = None

class FinalPerformanceReviewSubmit(BaseModel):
    overall_performance_rating: int
    overall_performance_comments: str

class CloseAlertRequest(BaseModel):
    new_deadline: Optional[str] = None

class GrievanceCreate(BaseModel):
    diary_kra_id: Optional[str] = None
    grievance_type: str  # kra | overall
    description: str

class GrievanceRespond(BaseModel):
    response: str
    resolve: bool = False

class BandCreate(BaseModel):
    band_code: str
    band_name: str

class KRAMasterCreate(BaseModel):
    band_id: Optional[str] = None
    kra_code: str
    kra_name: str
    kra_description: Optional[str] = None
    is_org_mandatory: bool = False

class KPIMasterCreate(BaseModel):
    kra_master_id: str
    kpi_code: str
    kpi_name: str
    kpi_description: Optional[str] = None

class KRAMandatoryUpdate(BaseModel):
    is_mandatory: bool

# ─── Auth ──────────────────────────────────────────────────────────────────────

@router.post("/auth/login")
def login(req: LoginRequest, db: Session = Depends(get_db)):
    username = req.username.strip()
    admin = db.query(AdminCredential).filter(AdminCredential.username == username, AdminCredential.is_active == True).first()
    if admin and verify_password(req.password, admin.password_hash):
        emp = db.query(Employee).filter(Employee.id == admin.employee_id).first()
        admin.last_login = datetime.utcnow()
        db.commit()
        token = _login_token_for(emp, req.remember_me)
        return {"access_token": token, "token_type": "bearer", "user": _emp_dict(emp), "must_change_password": False}

    if AUTH_MODE in {"password", "local"}:
        emp = db.query(Employee).filter(
            or_(
                func.lower(Employee.email) == username.lower(),
                func.lower(Employee.employee_code) == username.lower(),
            ),
            Employee.is_active == True,
        ).first()
        employee_credential = None
        if emp:
            employee_credential = db.query(EmployeeCredential).filter(
                EmployeeCredential.employee_id == emp.id,
                EmployeeCredential.is_active == True,
            ).first()
        if emp and employee_credential and verify_password(req.password, employee_credential.password_hash):
            employee_credential.last_login = datetime.utcnow()
            db.commit()
            user = _emp_dict(emp)
            user["must_change_password"] = employee_credential.must_change_password
            token = _login_token_for(emp, req.remember_me)
            return {
                "access_token": token,
                "token_type": "bearer",
                "user": user,
                "must_change_password": employee_credential.must_change_password,
            }
    raise HTTPException(status_code=401, detail="Email or password is incorrect")

@router.post("/auth/change-password")
def change_password(req: ChangePasswordRequest, db: Session = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    if AUTH_MODE not in {"password", "local"}:
        raise HTTPException(status_code=400, detail="EDGE password login is disabled for this environment")

    if req.confirm_email.strip().lower() != current_user.email.lower():
        raise HTTPException(status_code=400, detail="Confirm the logged-in EDGE email before changing password")

    credential = db.query(EmployeeCredential).filter(
        EmployeeCredential.employee_id == current_user.id,
        EmployeeCredential.is_active == True,
    ).first()
    if not credential or not verify_password(req.current_password, credential.password_hash):
        raise HTTPException(status_code=400, detail="Current password is incorrect")

    new_password = req.new_password.strip()
    if len(new_password) < 8:
        raise HTTPException(status_code=400, detail="New password must be at least 8 characters")
    if new_password == DEFAULT_EMPLOYEE_PASSWORD:
        raise HTTPException(status_code=400, detail="New password cannot be the default EDGE password")
    if verify_password(new_password, credential.password_hash):
        raise HTTPException(status_code=400, detail="New password must be different from current password")

    credential.password_hash = hash_password(new_password)
    credential.must_change_password = False
    credential.password_changed_at = datetime.utcnow()
    db.commit()
    return {"message": "Password updated successfully. Please log in with your new password."}


@router.post("/auth/forgot-password")
def forgot_password(req: ForgotPasswordRequest, db: Session = Depends(get_db)):
    email = req.email.strip().lower()
    generic_message = "If an active account exists for this email, a password reset link has been sent."
    if not email:
        raise HTTPException(status_code=400, detail="Email is required")

    emp = db.query(Employee).filter(func.lower(Employee.email) == email, Employee.is_active == True).first()
    credential = None
    if emp:
        credential = db.query(EmployeeCredential).filter(
            EmployeeCredential.employee_id == emp.id,
            EmployeeCredential.is_active == True,
        ).first()

    if not emp or not credential:
        return {"message": generic_message}

    raw_token = secrets.token_urlsafe(32)
    token_hash = hashlib.sha256(raw_token.encode()).hexdigest()
    expires_minutes = int(os.getenv("PASSWORD_RESET_EXPIRE_MINUTES", "10"))
    now = datetime.utcnow()

    db.query(PasswordResetToken).filter(
        PasswordResetToken.employee_credential_id == credential.id,
        PasswordResetToken.used_at.is_(None),
        PasswordResetToken.expires_at > now,
    ).update({"used_at": now}, synchronize_session=False)

    db.add(PasswordResetToken(
        id=uuid.uuid4(),
        employee_credential_id=credential.id,
        token_hash=token_hash,
        expires_at=now + timedelta(minutes=expires_minutes),
    ))
    db.commit()

    frontend_origin = os.getenv("FRONTEND_ORIGIN", "http://localhost:5173").rstrip("/")
    reset_link = f"{frontend_origin}/reset-password?token={raw_token}"
    try:
        send_password_reset_email(emp.email, reset_link, expires_minutes)
    except Exception as exc:
        print(f"Password reset email failed for {emp.email}: {exc}")
    return {"message": generic_message}


@router.post("/auth/reset-password")
def reset_password(req: ResetPasswordRequest, db: Session = Depends(get_db)):
    raw_token = req.token.strip()
    password = req.password.strip()
    if not raw_token:
        raise HTTPException(status_code=400, detail="Reset token is required")
    if len(password) < 8:
        raise HTTPException(status_code=400, detail="Password must be at least 8 characters")

    token_hash = hashlib.sha256(raw_token.encode()).hexdigest()
    now = datetime.utcnow()
    reset_token = db.query(PasswordResetToken).filter(
        PasswordResetToken.token_hash == token_hash,
        PasswordResetToken.used_at.is_(None),
        PasswordResetToken.expires_at > now,
    ).first()
    if not reset_token:
        raise HTTPException(status_code=400, detail="Reset link is invalid or expired")

    credential = db.query(EmployeeCredential).filter(
        EmployeeCredential.id == reset_token.employee_credential_id,
        EmployeeCredential.is_active == True,
    ).first()
    if not credential:
        raise HTTPException(status_code=400, detail="Reset link is invalid or expired")

    if password == DEFAULT_EMPLOYEE_PASSWORD:
        raise HTTPException(status_code=400, detail="New password cannot be the default EDGE password")

    credential.password_hash = hash_password(password)
    credential.must_change_password = False
    credential.password_changed_at = now
    credential.updated_at = now
    reset_token.used_at = now
    db.commit()
    return {"message": "Password has been reset successfully. Please log in with your new password."}


@router.post("/auth/microsoft")
def microsoft_login(req: MicrosoftLoginRequest, db: Session = Depends(get_db)):
    claims = validate_microsoft_id_token(req.id_token)
    emp = db.query(Employee).filter(
        func.lower(Employee.email) == claims["resolved_email"],
        Employee.is_active == True,
    ).first()
    if not emp:
        raise HTTPException(status_code=403, detail="This Microsoft account is not provisioned in PMS")

    token = create_access_token({"sub": str(emp.id), "role": emp.role.value, "name": emp.full_name})
    return {"access_token": token, "token_type": "bearer", "user": _emp_dict(emp)}

@router.get("/auth/me")
def me(current_user: Employee = Depends(get_current_user)):
    return _emp_dict(current_user)


def _matches_employee_search(emp: Optional[Employee], search: Optional[str]) -> bool:
    term = " ".join((search or "").lower().split())
    if not term:
        return True
    if not emp:
        return False

    full_name = " ".join((emp.full_name or "").lower().split())
    name_matches = term in full_name if " " in term else any(
        part.startswith(term) for part in full_name.split()
    )
    if name_matches:
        return True

    band = getattr(emp, "band", None)
    email_local = (emp.email or "").split("@", 1)[0]
    search_values = [
        email_local,
        emp.employee_code or "",
        emp.band_code or "",
        emp.band_name or "",
    ]
    if band:
        search_values.extend([band.band_code or "", band.band_name or ""])

    if len(term) <= 2:
        return any(str(value).lower().startswith(term) for value in search_values)

    return any(term in str(value).lower() for value in search_values)

# ─── Employees ─────────────────────────────────────────────────────────────────

@router.get("/employees")
def list_employees(
    page: int = 1,
    page_size: int = 25,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user),
):
    page = max(page, 1)
    page_size = min(max(page_size, 1), 100)
    q = db.query(Employee).options(joinedload(Employee.band)).filter(Employee.is_active == True)
    if search:
        term = f"%{search.strip()}%"
        q = q.filter(or_(
            Employee.full_name.ilike(term),
            Employee.email.ilike(term),
            Employee.employee_code.ilike(term),
            Employee.band_code.ilike(term),
            Employee.band_name.ilike(term),
        ))

    total = q.count()
    emps = q.order_by(Employee.full_name.asc()).offset((page - 1) * page_size).limit(page_size).all()
    manager_ids = {e.manager_id for e in emps if e.manager_id}
    managers = {
        m.id: m.full_name for m in db.query(Employee.id, Employee.full_name).filter(Employee.id.in_(manager_ids)).all()
    } if manager_ids else {}
    items = []
    for emp in emps:
        item = _emp_dict(emp)
        item["manager_name"] = managers.get(emp.manager_id)
        items.append(item)
    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size,
    }

@router.get("/employees/my-team")
def my_team(db: Session = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    team = db.query(Employee).options(joinedload(Employee.band)).filter(
        Employee.manager_id == current_user.id,
        Employee.is_active == True,
        Employee.id != current_user.id
    ).all()
    return [_emp_dict(e) for e in team]

@router.post("/employees")
def create_employee(data: EmployeeCreate, db: Session = Depends(get_db), current_user: Employee = Depends(require_role("admin"))):
    payload = data.dict()
    if not payload.get("manager_id"): payload["manager_id"] = None
    if not payload.get("approver_id"): payload["approver_id"] = None
    emp = Employee(id=str(uuid.uuid4()), **payload)
    db.add(emp)
    db.flush()
    if emp.role != RoleEnum.admin:
        db.add(EmployeeCredential(
            id=uuid.uuid4(),
            employee_id=emp.id,
            password_hash=hash_password(DEFAULT_EMPLOYEE_PASSWORD),
        ))
    db.commit()
    return _emp_dict(emp)

@router.post("/employees/sync-nest")
def sync_employees_from_nest(db: Session = Depends(get_db), _=Depends(require_role("admin"))):
    try:
        result = sync_nest_employees(db)
        ensure_employee_credentials(db)
        return {"message": "NEST employee sync completed", **result}
    except ValueError as exc:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(exc))
    except requests.RequestException as exc:
        db.rollback()
        raise HTTPException(status_code=502, detail=f"NEST employee API request failed: {exc}")

# ─── Bands ─────────────────────────────────────────────────────────────────────

@router.get("/bands")
def list_bands(db: Session = Depends(get_db), _=Depends(get_current_user)):
    return [{"id": b.id, "band_code": b.band_code, "band_name": b.band_name} for b in db.query(Band).filter(Band.is_active == True).all()]

@router.post("/bands")
def create_band(data: BandCreate, db: Session = Depends(get_db), _=Depends(require_role("admin"))):
    b = Band(id=str(uuid.uuid4()), **data.dict())
    db.add(b); db.commit()
    return {"id": b.id, "band_code": b.band_code, "band_name": b.band_name}

# ─── KRA Master ────────────────────────────────────────────────────────────────

@router.get("/kra-master")
def list_kra_master(band_id: Optional[str] = None, db: Session = Depends(get_db), _=Depends(get_current_user)):
    q = db.query(KRAMaster).filter(KRAMaster.is_active == True)
    if band_id:
        q = q.filter(or_(KRAMaster.band_id == band_id, KRAMaster.band_id.is_(None)))
    kras = q.options(joinedload(KRAMaster.kpi_masters)).all()
    return [{"id": k.id, "kra_code": k.kra_code, "kra_name": k.kra_name, "kra_description": k.kra_description,
             "band_id": k.band_id, "is_mandatory": k.is_mandatory, "is_org_mandatory": k.is_org_mandatory,
             "scope": "org" if k.band_id is None else "band",
             "kpis": [{"id": p.id, "kpi_code": p.kpi_code, "kpi_name": p.kpi_name, "kpi_description": p.kpi_description} for p in k.kpi_masters if p.is_active]} for k in kras]

@router.post("/kra-master")
def create_kra(data: KRAMasterCreate, db: Session = Depends(get_db), _=Depends(require_role("admin"))):
    if data.band_id:
        raise HTTPException(400, "Admin can create only org-level KRAs from this form")
    payload = data.dict()
    payload["band_id"] = None
    k = KRAMaster(id=str(uuid.uuid4()), **payload); db.add(k); db.commit()
    return {"id": k.id, "kra_name": k.kra_name}

@router.delete("/kra-master/{kra_id}")
def delete_kra(kra_id: str, db: Session = Depends(get_db), _=Depends(require_role("admin"))):
    k = _get_or_404(db, KRAMaster, kra_id)
    k.is_active = False
    for kpi in db.query(KPIMaster).filter(KPIMaster.kra_master_id == k.id).all():
        kpi.is_active = False
    db.commit()
    return {"message": "KRA deleted"}

@router.put("/kra-master/{kra_id}/mandatory")
def update_kra_mandatory(kra_id: str, data: KRAMandatoryUpdate, db: Session = Depends(get_db), _=Depends(require_role("admin"))):
    k = _get_or_404(db, KRAMaster, kra_id)
    k.is_mandatory = data.is_mandatory
    db.commit()
    return {"id": k.id, "is_mandatory": k.is_mandatory}

@router.post("/kpi-master")
def create_kpi(data: KPIMasterCreate, db: Session = Depends(get_db), _=Depends(require_role("admin"))):
    k = KPIMaster(id=str(uuid.uuid4()), **data.dict()); db.add(k); db.commit()
    return {"id": k.id, "kpi_name": k.kpi_name}

@router.post("/kra-master/import-excel")
async def import_kra_excel(file: UploadFile = File(...), db: Session = Depends(get_db), _=Depends(require_role("admin"))):
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "Please upload an Excel .xlsx or .xlsm file")

    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(500, "Excel import dependency is missing. Install openpyxl.")

    from io import BytesIO

    content = await file.read()
    try:
        workbook = load_workbook(BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(400, "Unable to read Excel file")

    bands = {_normalize_band_code(band.band_code): band for band in db.query(Band).filter(Band.is_active == True).all()}
    imported = {"sheets": 0, "kras": 0, "kpis": 0, "skipped_sheets": []}

    for sheet in workbook.worksheets:
        band = bands.get(_normalize_band_code(sheet.title))
        if not band:
            imported["skipped_sheets"].append({"sheet": sheet.title, "reason": "No matching band code"})
            continue

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        header_info = _find_excel_header_row(rows)
        if not header_info:
            imported["skipped_sheets"].append({"sheet": sheet.title, "reason": "No KRA / Key KPIs header row found"})
            continue
        header_index, headers = header_info

        imported["sheets"] += 1
        current_kra = None
        kra_sequence = 0
        kpi_sequences = {}
        sheet_seed = _excel_code_seed(sheet.title)

        bad_kras = db.query(KRAMaster).filter(
            KRAMaster.band_id == band.id,
            func.lower(KRAMaster.kra_name) == "kra",
        ).all()
        for bad_kra in bad_kras:
            bad_kra.is_active = False
            for bad_kpi in db.query(KPIMaster).filter(KPIMaster.kra_master_id == bad_kra.id).all():
                bad_kpi.is_active = False

        for values in rows[header_index + 1:]:
            row = {
                header: _cell_text(values[index]) if index < len(values) else None
                for index, header in enumerate(headers)
                if header
            }

            raw_kra_code = _get_excel_value(row, "KRA Code", "KRA ID")
            kra_name = _get_excel_value(row, "KRA Name", "KRA")
            kra_description = _get_excel_value(row, "KRA Description", "KRA Desc", "Description")
            raw_kpi_code = _get_excel_value(row, "KPI Code", "KPI ID")
            kpi_names = _split_key_kpis(_get_excel_value(row, "Key KPIs", "Key KPI", "KPI Name", "KPI"))
            kpi_description = _get_excel_value(row, "KPI Description", "KPI Desc")

            if _normalize_excel_header(kra_name) == "kra" and any(_normalize_excel_header(kpi) in {"keykpi", "keykpis", "kpi", "kpiname"} for kpi in kpi_names):
                current_kra = None
                continue

            if kra_name:
                kra_sequence += 1
                kra_code = _clean_excel_code(raw_kra_code, f"{sheet_seed}-KRA{kra_sequence:03d}")
                current_kra = None
                if raw_kra_code and len(raw_kra_code) <= 20:
                    current_kra = db.query(KRAMaster).filter(
                        KRAMaster.band_id == band.id,
                        func.lower(KRAMaster.kra_code) == kra_code.lower(),
                    ).first()
                if not current_kra:
                    current_kra = db.query(KRAMaster).filter(
                        KRAMaster.band_id == band.id,
                        func.lower(KRAMaster.kra_name) == kra_name.lower(),
                    ).first()
                if current_kra:
                    current_kra.kra_code = kra_code
                    current_kra.kra_name = kra_name
                    current_kra.kra_description = kra_description
                    current_kra.is_active = True
                else:
                    current_kra = KRAMaster(
                        id=str(uuid.uuid4()),
                        band_id=band.id,
                        kra_code=kra_code,
                        kra_name=kra_name,
                        kra_description=kra_description,
                    )
                    db.add(current_kra)
                    db.flush()
                    imported["kras"] += 1

            if not current_kra or not kpi_names:
                continue

            for kpi_name in kpi_names:
                kpi_count = kpi_sequences.get(current_kra.id, 0) + 1
                kpi_sequences[current_kra.id] = kpi_count
                kpi_code = _clean_excel_code(raw_kpi_code if len(kpi_names) == 1 else None, f"{current_kra.kra_code}-K{kpi_count:02d}")

                kpi = db.query(KPIMaster).filter(
                    func.lower(KPIMaster.kpi_name) == kpi_name.lower(),
                    KPIMaster.kra_master_id == current_kra.id,
                ).first()
                if kpi:
                    kpi.kpi_code = kpi_code
                    kpi.kpi_name = kpi_name
                    kpi.kpi_description = kpi_description
                    kpi.is_active = True
                else:
                    db.add(KPIMaster(
                        id=str(uuid.uuid4()),
                        kra_master_id=current_kra.id,
                        kpi_code=kpi_code,
                        kpi_name=kpi_name,
                        kpi_description=kpi_description,
                    ))
                    imported["kpis"] += 1

    db.commit()
    return imported

@router.delete("/kpi-master/{kpi_id}")
def delete_kpi(kpi_id: str, db: Session = Depends(get_db), _=Depends(require_role("admin"))):
    kpi = _get_or_404(db, KPIMaster, kpi_id)
    kpi.is_active = False
    db.commit()
    return {"message": "KPI deleted"}

# ─── Performance Cycles ────────────────────────────────────────────────────────

@router.get("/cycles")
def list_cycles(db: Session = Depends(get_db), _=Depends(get_current_user)):
    cycles = db.query(PerformanceCycle).order_by(PerformanceCycle.created_at.desc()).all()
    return [_cycle_dict(c) for c in cycles]

@router.get("/cycles/active")
def active_cycle(db: Session = Depends(get_db), _=Depends(get_current_user)):
    c = db.query(PerformanceCycle).filter(PerformanceCycle.status.in_(["step1", "step2"])).first()
    return _cycle_dict(c) if c else None

@router.post("/cycles")
def create_cycle(data: CycleCreate, db: Session = Depends(get_db), current_user: Employee = Depends(require_role("admin"))):
    if not data.cycle_name or not data.cycle_name.strip():
        raise HTTPException(400, "Cycle name is required")
    def pd(s): return datetime.strptime(s, "%Y-%m-%d")
    db_period = PUBLIC_TO_DB_CYCLE_PERIOD.get(data.period)
    if not db_period:
        raise HTTPException(400, "Invalid cycle period. Use H1_SEP or H2_MAR.")
    c = PerformanceCycle(
        id=str(uuid.uuid4()), cycle_name=data.cycle_name, financial_year=data.financial_year,
        period=db_period, status="draft",
        step1_open_date=pd(data.step1_open_date), step1_kra_deadline=pd(data.step1_kra_deadline),
        step1_approval_date=pd(data.step1_approval_date), created_by=current_user.id
    )
    db.add(c); db.commit()
    _create_diaries_for_cycle(c.id, db)
    return _cycle_dict(c)

@router.post("/cycles/{cycle_id}/activate")
def activate_cycle(cycle_id: str, db: Session = Depends(get_db), _=Depends(require_role("admin"))):
    c = _get_or_404(db, PerformanceCycle, cycle_id)
    c.status = CycleStatus.step1
    diaries = db.query(PerformanceDiary).filter(PerformanceDiary.cycle_id == cycle_id).all()
    notifications = []
    recipients = set()
    for d in diaries:
        recipients.add(d.employee_id)
        recipients.add(d.manager_id)
        recipients.add(d.approver_id)

    recipients = _active_employee_recipient_ids(db, recipients)
    for rid in recipients:
        notifications.append(Notification(
            id=str(uuid.uuid4()), recipient_id=rid, event_type="step1_activated",
            title="Performance Cycle Initiated - KRAs Allocation",
            body=_step1_activated_body(c.cycle_name, c.period, c.step1_approval_date.strftime("%Y-%m-%d")),
            channel=NotifChannel.both
        ))
    db.add_all(notifications)
    db.commit()
    _send_emails_for_notifications_async(db, notifications)
    return {"message": "Cycle activated and alerts sent to reportees, managers, and approvers"}

@router.post("/cycles/{cycle_id}/send-step1-alert")
def send_step1_alert(cycle_id: str, db: Session = Depends(get_db), _=Depends(require_role("admin"))):
    c = _get_or_404(db, PerformanceCycle, cycle_id)
    diaries = db.query(PerformanceDiary).filter(PerformanceDiary.cycle_id == cycle_id).all()
    recipients = set()
    for d in diaries:
        recipients.add(d.employee_id)
        recipients.add(d.manager_id)
        recipients.add(d.approver_id)

    recipients = _active_employee_recipient_ids(db, recipients)
    notifications = []
    for rid in recipients:
        notifications.append(Notification(
            id=str(uuid.uuid4()), recipient_id=rid, event_type="step1_deadline_alert",
            title="KRA Allocation Deadline Reminder",
            body=_step1_deadline_alert_body(c.cycle_name, c.period, c.step1_approval_date.strftime("%Y-%m-%d")),
            channel=NotifChannel.both
        ))
    db.add_all(notifications)
    db.commit()
    _send_emails_for_notifications_async(db, notifications)
    return {"message": "KRA Allocation reminder alerts sent to reportees, managers, and approvers"}

@router.put("/cycles/{cycle_id}/step1-approval-date")
def update_step1_approval_date(
    cycle_id: str,
    data: Step1ApprovalDateUpdate,
    db: Session = Depends(get_db),
    _=Depends(require_role("admin")),
):
    c = _get_or_404(db, PerformanceCycle, cycle_id)
    if c.status == CycleStatus.closed:
        raise HTTPException(400, "Step 1 end date cannot be changed for a closed cycle")
    try:
        approval_date = datetime.strptime(data.step1_approval_date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(400, "Invalid date format. Use YYYY-MM-DD")
    if approval_date < c.step1_kra_deadline:
        raise HTTPException(400, "Step 1 end date cannot be before the KRA submission deadline")
    if c.step2_open_date and approval_date > c.step2_open_date:
        raise HTTPException(400, "Step 1 end date cannot be after the Step 2 opening date")

    c.step1_approval_date = approval_date
    db.commit()
    return {
        "message": "Step 1 end date updated",
        "step1_approval_date": c.step1_approval_date.strftime("%Y-%m-%d"),
    }

@router.post("/cycles/{cycle_id}/open-step2")
def open_step2(cycle_id: str, data: Step2OpenRequest, db: Session = Depends(get_db), _=Depends(require_role("admin"))):
    c = _get_or_404(db, PerformanceCycle, cycle_id)
    if c.status not in [CycleStatus.step1, "step1"]:
        raise HTTPException(400, "Rating can be opened only for an active KRA Allocation cycle")
    def pd(s): return datetime.strptime(s, "%Y-%m-%d")
    c.step2_open_date = pd(data.step2_open_date)
    c.step2_self_deadline = pd(data.step2_self_deadline)
    c.step2_mgr_deadline = pd(data.step2_mgr_deadline)
    c.step2_approval_date = pd(data.step2_approval_date)
    c.status = CycleStatus.step2
    diaries = db.query(PerformanceDiary).filter(PerformanceDiary.cycle_id == cycle_id).all()
    notifications = []
    recipients = set()
    for d in diaries:
        d.self_status = SelfStatus.open if d.self_status == SelfStatus.not_open else d.self_status
        recipients.add(d.employee_id)
        recipients.add(d.manager_id)
        recipients.add(d.approver_id)

    recipients = _active_employee_recipient_ids(db, recipients)
    for rid in recipients:
        notifications.append(Notification(
            id=str(uuid.uuid4()), recipient_id=rid, event_type="step2_activated",
            title="Performance Cycle - Rating Open",
            body=_step2_activated_body(c.cycle_name, c.period, c.step2_self_deadline.strftime("%Y-%m-%d"), c.step2_mgr_deadline.strftime("%Y-%m-%d")),
            channel=NotifChannel.both
        ))
    db.add_all(notifications)
    db.commit()
    _send_emails_for_notifications_async(db, notifications)
    return {"message": "Rating opened and alerts sent to all"}

@router.post("/cycles/{cycle_id}/send-step2-alert")
def send_step2_alert(cycle_id: str, db: Session = Depends(get_db), _=Depends(require_role("admin"))):
    c = _get_or_404(db, PerformanceCycle, cycle_id)
    diaries = db.query(PerformanceDiary).filter(PerformanceDiary.cycle_id == cycle_id).all()
    recipients = set()
    for d in diaries:
        recipients.add(d.employee_id)
        recipients.add(d.manager_id)
        recipients.add(d.approver_id)

    recipients = _active_employee_recipient_ids(db, recipients)
    notifications = []
    for rid in recipients:
        notifications.append(Notification(
            id=str(uuid.uuid4()), recipient_id=rid, event_type="step2_deadline_alert",
            title="Performance Rating Deadline Reminder",
            body=_step2_deadline_alert_body(c.cycle_name, c.period, c.step2_self_deadline.strftime("%Y-%m-%d"), c.step2_mgr_deadline.strftime("%Y-%m-%d")),
            channel=NotifChannel.both
        ))
    db.add_all(notifications)
    db.commit()
    _send_emails_for_notifications_async(db, notifications)
    return {"message": "Rating reminder alerts sent"}

@router.post("/cycles/{cycle_id}/close")
def close_cycle(cycle_id: str, db: Session = Depends(get_db), _=Depends(require_role("admin"))):
    c = _get_or_404(db, PerformanceCycle, cycle_id)
    diaries = db.query(PerformanceDiary).filter(PerformanceDiary.cycle_id == cycle_id).all()
    if not diaries:
        raise HTTPException(400, "No diaries found for this cycle")

    c.status = CycleStatus.closed
    employee_recipients = set()
    other_recipients = set()
    for d in diaries:
        d.final_status = FinalStatus.closed
        d.self_status = SelfStatus.submitted if d.self_status == SelfStatus.open else d.self_status
        employee_recipients.add(d.employee_id)
        other_recipients.update(rid for rid in [d.manager_id, d.approver_id] if rid)

    employee_recipients = _active_employee_recipient_ids(db, employee_recipients)
    other_recipients = _active_employee_recipient_ids(db, other_recipients)
    other_recipients -= employee_recipients
    notifications = [
        Notification(
            id=str(uuid.uuid4()),
            recipient_id=recipient_id,
            diary_id=None,
            event_type="cycle_closed",
            title="Performance cycle closed",
            body=_cycle_closed_body(c.cycle_name, c.period, is_employee=True),
            channel=NotifChannel.both,
        )
        for recipient_id in employee_recipients
    ] + [
        Notification(
            id=str(uuid.uuid4()),
            recipient_id=recipient_id,
            diary_id=None,
            event_type="cycle_closed",
            title="Performance cycle closed",
            body=_cycle_closed_body(c.cycle_name, c.period),
            channel=NotifChannel.both,
        )
        for recipient_id in other_recipients
    ]
    db.add_all(notifications)
    db.commit()
    _send_emails_for_notifications_async(db, notifications)
    return {"message": "Cycle closed and alerts sent"}

@router.post("/cycles/{cycle_id}/send-close-alert")
def send_close_alert(cycle_id: str, data: CloseAlertRequest, db: Session = Depends(get_db), _=Depends(require_role("admin"))):
    c = _get_or_404(db, PerformanceCycle, cycle_id)
    if data.new_deadline:
        try:
            c.step2_approval_date = datetime.strptime(data.new_deadline, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(400, "Invalid date format. Use YYYY-MM-DD")
    
    diaries = db.query(PerformanceDiary).filter(PerformanceDiary.cycle_id == cycle_id).all()
    if not diaries:
        raise HTTPException(400, "No diaries found for this cycle")

    deadline_str = c.step2_approval_date.strftime("%Y-%m-%d")
    body = _cycle_close_alert_body(c.cycle_name, c.period, deadline_str)
    
    recipient_ids = set()
    for d in diaries:
        recipient_ids.update(rid for rid in [d.employee_id, d.manager_id, d.approver_id] if rid)

    recipient_ids = _active_employee_recipient_ids(db, recipient_ids)
    notifications = [
        Notification(
            id=str(uuid.uuid4()),
            recipient_id=recipient_id,
            diary_id=None,
            event_type="cycle_close_alert",
            title="Cycle closure alert",
            body=body,
            channel=NotifChannel.both,
        )
        for recipient_id in recipient_ids
    ]
    db.add_all(notifications)
    db.commit()
    _send_emails_for_notifications_async(db, notifications)
    return {"message": "Cycle close alerts sent"}

# ─── Performance Diaries ───────────────────────────────────────────────────────

@router.get("/diaries/my")
def my_diaries(page: Optional[int] = None, page_size: int = 25, db: Session = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    diaries = db.query(PerformanceDiary).filter(PerformanceDiary.employee_id == current_user.id).options(joinedload(PerformanceDiary.cycle)).all()
    open_grievance_ids = _open_grievance_diary_ids(db, [d.id for d in diaries])
    result = []
    for d in diaries:
        dd = _diary_dict(d)
        dd["has_open_grievance"] = d.id in open_grievance_ids
        result.append(dd)
    if page:
        page = max(page, 1)
        page_size = min(max(page_size, 1), 100)
        total = len(result)
        start = (page - 1) * page_size
        return {"items": result[start:start + page_size], "total": total, "page": page, "page_size": page_size, "total_pages": (total + page_size - 1) // page_size}
    return result

def _employee_name_sort_value(item):
    employee = item.get("employee") or {}
    return (employee.get("full_name") or "").lower()

def _status_changed(item):
    return (
        item.get("kra_status") != "draft"
        or item.get("self_status") != "not_open"
        or item.get("mgr_status") != "pending"
        or item.get("final_status") != "pending"
        or bool(item.get("final_review_open"))
        or bool(item.get("has_open_grievance"))
        or (item.get("kra_sendback_count") or 0) > 0
        or (item.get("mgr_sendback_count") or 0) > 0
    )

def _sort_time_value(item):
    value = item.get("updated_at") or item.get("created_at")
    if not value:
        return 0
    try:
        return datetime.fromisoformat(value).timestamp()
    except ValueError:
        return 0

def _status_update_sort_key(item):
    return (0 if _status_changed(item) else 1, -_sort_time_value(item), _employee_name_sort_value(item))

def _diary_status_counts(items):
    approver_rating_done = sum(
        1 for item in items
        if bool(item.get("final_review_open")) or item.get("final_status") in ["baselined", "closed"]
    )
    return {
        "total": len(items),
        "kra_pending": sum(1 for item in items if item.get("kra_status") != "baselined"),
        "kra_pending_with_manager": sum(1 for item in items if item.get("kra_status") in ["draft", "sent_back"]),
        "kra_baselined": sum(1 for item in items if item.get("kra_status") == "baselined"),
        "kra_submitted": sum(1 for item in items if item.get("kra_status") == "submitted"),
        "kra_sent_back": sum(1 for item in items if item.get("kra_status") == "sent_back"),
        "sent_back": sum(1 for item in items if item.get("kra_status") == "sent_back" or item.get("mgr_status") == "sent_back"),
        "self_rating_done": sum(1 for item in items if item.get("self_status") in ["submitted", "auto_submitted"]),
        "manager_rating_pending": sum(
            1 for item in items
            if item.get("final_status") not in ["baselined", "closed"]
            and item.get("mgr_status") != "submitted"
            and (
                item.get("mgr_status") == "sent_back"
                or item.get("self_status") in ["submitted", "auto_submitted"]
            )
        ),
        "manager_rating_done": sum(1 for item in items if item.get("mgr_status") == "submitted"),
        "approver_rating_done": approver_rating_done,
        "rating_baselined": approver_rating_done,
        "final_review_done": sum(1 for item in items if item.get("final_status") in ["baselined", "closed"]),
        "ratings_pending": sum(
            1 for item in items
            if item.get("mgr_status") == "submitted"
            and not item.get("final_review_open")
            and item.get("final_status") not in ["baselined", "closed"]
        ),
        "completed": sum(1 for item in items if item.get("final_status") in ["baselined", "closed"]),
    }

@router.get("/diaries/team")
def team_diaries(cycle_id: Optional[str] = None, status_filter: Optional[str] = None, search: Optional[str] = None, page: Optional[int] = None, page_size: int = 25, db: Session = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    q = db.query(PerformanceDiary).options(joinedload(PerformanceDiary.cycle))
    if current_user.role.value == "admin":
        q = q.join(Employee, PerformanceDiary.employee_id == Employee.id).filter(
            PerformanceDiary.employee_id != current_user.id,
            Employee.is_active == True,
        )
    else:
        q = q.join(Employee, PerformanceDiary.employee_id == Employee.id).filter(
            PerformanceDiary.manager_id == current_user.id,
            PerformanceDiary.employee_id != current_user.id,
            Employee.is_active == True,
        )
    if cycle_id: q = q.filter(PerformanceDiary.cycle_id == cycle_id)
    diaries = q.all()
    employee_ids = {d.employee_id for d in diaries}
    manager_ids = {d.manager_id for d in diaries if d.manager_id}
    approver_ids = {d.approver_id for d in diaries if d.approver_id}
    people_ids = employee_ids | manager_ids | approver_ids
    employees = {
        e.id: e for e in db.query(Employee).options(joinedload(Employee.band)).filter(Employee.id.in_(people_ids)).all()
    } if people_ids else {}
    open_grievance_ids = _open_grievance_diary_ids(db, [d.id for d in diaries])
    result = []
    for d in diaries:
        emp = employees.get(d.employee_id)
        manager = employees.get(d.manager_id) if d.manager_id else None
        approver = employees.get(d.approver_id) if d.approver_id else None
        dd = _diary_dict(d)
        dd["has_open_grievance"] = d.id in open_grievance_ids
        dd["employee"] = _emp_dict(emp) if emp else None
        dd["manager"] = _emp_dict(manager)
        dd["approver"] = _emp_dict(approver)
        if search:
            searchable_people = [emp]
            if current_user.role.value == "admin":
                searchable_people.extend([manager, approver])
            if not any(_matches_employee_search(person, search) for person in searchable_people):
                continue
        if status_filter:
            s = status_filter.lower()
            is_completed = d.final_status.value in ["baselined", "closed"]
            is_rating_started = d.self_status.value != "not_open" or d.mgr_status.value != "pending" or d.final_review_open
            is_step2 = not is_completed and d.kra_status == DiaryKRAStatus.baselined and is_rating_started
            is_step1 = not is_completed and not is_step2
            is_self_rating = (
                not is_completed
                and d.mgr_status == MgrStatus.pending
                and d.self_status in [SelfStatus.not_open, SelfStatus.open]
                and (d.cycle.status == CycleStatus.step2 or d.self_status == SelfStatus.open)
                and not d.final_review_open
            )

            if s == "not_started" and d.kra_status.value != "draft": continue
            elif s == "in_draft" and d.kra_status.value != "draft": continue
            elif s == "submitted" and d.kra_status.value != "submitted": continue
            elif s == "sent_back" and d.kra_status.value != "sent_back": continue
            elif s == "baselined" and d.kra_status.value != "baselined": continue
            elif s == "self_rating" and not is_self_rating: continue
            elif s == "step1" and not is_step1: continue
            elif s == "step2" and not is_step2: continue
            elif s == "completed" and not is_completed: continue
            elif s == "pending" and d.final_status.value not in ["pending"]: continue
            elif s == "kra_allocation" and d.kra_status.value != "draft": continue
        result.append(dd)
    if current_user.role.value == "admin":
        result.sort(key=_employee_name_sort_value)
    else:
        result.sort(key=_status_update_sort_key)
    if page:
        page = max(page, 1)
        page_size = min(max(page_size, 1), 100)
        total = len(result)
        start = (page - 1) * page_size
        return {"items": result[start:start + page_size], "total": total, "page": page, "page_size": page_size, "total_pages": (total + page_size - 1) // page_size, "status_counts": _diary_status_counts(result)}
    return result

@router.get("/diaries/approver-queue")
def approver_queue(cycle_id: Optional[str] = None, status_filter: Optional[str] = None, search: Optional[str] = None, page: Optional[int] = None, page_size: int = 25, db: Session = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    q = db.query(PerformanceDiary).join(
        Employee,
        PerformanceDiary.employee_id == Employee.id,
    ).filter(
        PerformanceDiary.approver_id == current_user.id,
        Employee.is_active == True,
    )
    if cycle_id: q = q.filter(PerformanceDiary.cycle_id == cycle_id)
    diaries = q.all()
    employee_ids = {d.employee_id for d in diaries}
    manager_ids = {d.manager_id for d in diaries}
    people_ids = employee_ids | manager_ids
    people = {
        e.id: e for e in db.query(Employee).options(joinedload(Employee.band)).filter(Employee.id.in_(people_ids)).all()
    } if people_ids else {}
    open_grievance_ids = _open_grievance_diary_ids(db, [d.id for d in diaries])
    result = []
    for d in diaries:
        emp = people.get(d.employee_id)
        mgr = people.get(d.manager_id)
        if search and not _matches_employee_search(emp, search):
            continue
        if status_filter:
            s = status_filter.lower()
            is_completed = d.final_status.value in ["baselined", "closed"]
            is_rating_started = d.self_status.value != "not_open" or d.mgr_status.value != "pending" or d.final_review_open
            is_step2 = not is_completed and d.kra_status == DiaryKRAStatus.baselined and is_rating_started
            is_step1 = not is_completed and not is_step2
            is_self_rating = (
                not is_completed
                and d.mgr_status == MgrStatus.pending
                and d.self_status in [SelfStatus.not_open, SelfStatus.open]
                and (d.cycle.status == CycleStatus.step2 or d.self_status == SelfStatus.open)
                and not d.final_review_open
            )

            if s == "not_started" and d.kra_status.value != "draft": continue
            elif s == "in_draft" and d.kra_status.value != "draft": continue
            elif s == "submitted" and d.kra_status.value != "submitted": continue
            elif s == "sent_back" and d.kra_status.value != "sent_back": continue
            elif s == "baselined" and d.kra_status.value != "baselined": continue
            elif s == "self_rating" and not is_self_rating: continue
            elif s == "step1" and not is_step1: continue
            elif s == "step2" and not is_step2: continue
            elif s == "completed" and not is_completed: continue
            elif s == "pending" and d.final_status.value not in ["pending"]: continue
        dd = _diary_dict(d)
        dd["has_open_grievance"] = d.id in open_grievance_ids
        dd["employee"] = _emp_dict(emp) if emp else None
        dd["manager"] = _emp_dict(mgr) if mgr else None
        result.append(dd)
    result.sort(key=_status_update_sort_key)
    if page:
        page = max(page, 1)
        page_size = min(max(page_size, 1), 100)
        total = len(result)
        start = (page - 1) * page_size
        return {"items": result[start:start + page_size], "total": total, "page": page, "page_size": page_size, "total_pages": (total + page_size - 1) // page_size, "status_counts": _diary_status_counts(result)}
    return result

@router.get("/diaries/{diary_id}")
def get_diary(diary_id: str, db: Session = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    d = db.query(PerformanceDiary).filter(PerformanceDiary.id == diary_id).options(
        joinedload(PerformanceDiary.kras).joinedload(DiaryKRA.kpis).joinedload(DiaryKPI.kpi_master),
        joinedload(PerformanceDiary.kras).joinedload(DiaryKRA.kra_master),
        joinedload(PerformanceDiary.cycle)
    ).first()
    if not d: raise HTTPException(404, "Diary not found")
    if current_user.role != RoleEnum.admin and current_user.id not in [d.employee_id, d.manager_id, d.approver_id]:
        raise HTTPException(403, "Not authorised to view this diary")
    result = _diary_dict(d, db)
    emp = db.query(Employee).filter(Employee.id == d.employee_id).first()
    result["employee"] = _emp_dict(emp) if emp else None
    result["kras"] = []
    for kra in sorted(d.kras, key=lambda x: x.sort_order):
        kra_name = kra.custom_kra_name or (kra.kra_master.kra_name if kra.kra_master else "Custom KRA")
        kra_description = kra.custom_kra_description or (kra.kra_master.kra_description if kra.kra_master else None)
        kd = {"id": kra.id, "kra_master_id": kra.kra_master_id, "kra_name": kra_name,
              "kra_description": kra_description,
              "custom_kra_name": kra.custom_kra_name, "custom_kra_description": kra.custom_kra_description,
              "is_custom": bool(kra.custom_kra_name),
              "weightage_pct": float(kra.weightage_pct), "self_rating": kra.self_rating,
              "self_comments": kra.self_comments, "mgr_rating": kra.mgr_rating, "mgr_comments": kra.mgr_comments,
              "is_mandatory": bool(kra.kra_master.is_mandatory) if kra.kra_master else False,
              "kpis": [{"id": kp.id, "kpi_master_id": kp.kpi_master_id,
                        "custom_kpi_name": kp.custom_kpi_name,
                        "kpi_name": kp.custom_kpi_name or (kp.kpi_master.kpi_name if kp.kpi_master else "Custom KPI"),
                        "is_custom": bool(kp.custom_kpi_name),
                        "measurement_comment": kp.measurement_comment} for kp in kra.kpis]}
        result["kras"].append(kd)
    is_own_performance = current_user.role != RoleEnum.admin and d.employee_id == current_user.id
    is_cycle_closed = d.final_status == FinalStatus.closed or (d.cycle and d.cycle.status == CycleStatus.closed)
    if is_own_performance and not is_cycle_closed:
        result["overall_performance_rating"] = None
        result["overall_performance_comments"] = None
        for kra in result["kras"]:
            kra["mgr_rating"] = None
            kra["mgr_comments"] = None
    return result

# ─── KRA Allocation (Manager - Feature 2) ──────────────────────────────────────

@router.post("/diaries/{diary_id}/allocate-kras")
def allocate_kras(diary_id: str, data: KRAAllocationSubmit, draft: bool = False, db: Session = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    d = _get_or_404(db, PerformanceDiary, diary_id)
    is_admin = current_user.role == RoleEnum.admin
    if d.manager_id != current_user.id and not is_admin:
        raise HTTPException(403, "Not the manager for this diary")
    if d.kra_status not in [DiaryKRAStatus.draft, DiaryKRAStatus.sent_back, DiaryKRAStatus.submitted]:
        raise HTTPException(400, "KRAs can be updated only before baseline")
    _validate_kra_allocation_payload(db, data.kras, require_complete=not draft)
    db.query(Grievance).filter(Grievance.diary_id == diary_id).update({"diary_kra_id": None})
    existing_kra_ids = [row[0] for row in db.query(DiaryKRA.id).filter(DiaryKRA.diary_id == diary_id).all()]
    if existing_kra_ids:
        db.query(DiaryKPI).filter(DiaryKPI.diary_kra_id.in_(existing_kra_ids)).delete(synchronize_session=False)
    db.query(DiaryKRA).filter(DiaryKRA.diary_id == diary_id).delete(synchronize_session=False)
    for i, kra_data in enumerate(data.kras):
        kra = DiaryKRA(
            id=str(uuid.uuid4()),
            diary_id=diary_id,
            kra_master_id=kra_data.kra_master_id,
            custom_kra_name=kra_data.custom_kra_name,
            custom_kra_description=kra_data.custom_kra_description,
            weightage_pct=kra_data.weightage_pct,
            sort_order=i,
        )
        db.add(kra); db.flush()
        for kpi_id in kra_data.kpi_ids:
            kpi = DiaryKPI(
                id=str(uuid.uuid4()),
                diary_kra_id=kra.id,
                kpi_master_id=kpi_id,
                measurement_comment=kra_data.measurement_comment,
            )
            db.add(kpi)
        for custom_kpi in kra_data.custom_kpis:
            kpi = DiaryKPI(
                id=str(uuid.uuid4()),
                diary_kra_id=kra.id,
                custom_kpi_name=custom_kpi.custom_kpi_name,
                measurement_comment=kra_data.measurement_comment,
            )
            db.add(kpi)
    d.updated_at = datetime.utcnow()
    db.commit()
    return {"message": "KRAs saved"}

@router.post("/diaries/{diary_id}/submit-kras")
def submit_kras(diary_id: str, db: Session = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    d = _get_or_404(db, PerformanceDiary, diary_id)
    is_admin = current_user.role == RoleEnum.admin
    if d.manager_id != current_user.id and not is_admin:
        raise HTTPException(403, "Not authorised")
    if d.kra_status not in [DiaryKRAStatus.draft, DiaryKRAStatus.sent_back, DiaryKRAStatus.submitted]:
        raise HTTPException(400, "KRAs can be submitted only before baseline")
    kras = db.query(DiaryKRA).filter(DiaryKRA.diary_id == diary_id).all()
    if not kras: raise HTTPException(400, "No KRAs allocated")
    total = sum(float(k.weightage_pct) for k in kras)
    if abs(total - 100.0) > 0.01: raise HTTPException(400, "Weightage must total 100%")
    custom_total = sum(float(k.weightage_pct) for k in kras if k.custom_kra_name)
    if custom_total > 25.0:
        raise HTTPException(400, "Custom KRA weightage cannot exceed 25%")
    custom_kra_count = sum(1 for k in kras if k.custom_kra_name)
    if custom_kra_count > 1:
        raise HTTPException(400, "Only one custom KRA is allowed per employee")
    standard_kra_ids = [str(k.kra_master_id) for k in kras if k.kra_master_id]
    if len(set(standard_kra_ids)) != len(standard_kra_ids):
        raise HTTPException(400, "Duplicate KRAs are not allowed")
    for kra in kras:
        if float(kra.weightage_pct) <= 0:
            raise HTTPException(400, "Weightage must be greater than 0")
        if kra.custom_kra_name:
            if kra.kra_master_id:
                raise HTTPException(400, "Custom KRAs cannot reference a master KRA")
            if not kra.kpis:
                raise HTTPException(400, "At least one custom KPI is required for a custom KRA")
        else:
            if not kra.kra_master_id:
                raise HTTPException(400, "KRA selection is required")
            if not kra.kpis:
                raise HTTPException(400, "At least one KPI is required for every KRA")
        if any(not (kpi.measurement_comment or "").strip() for kpi in kra.kpis):
            raise HTTPException(400, "Measurement description is mandatory for every KRA")
    d.kra_status = DiaryKRAStatus.submitted
    reportee = db.query(Employee).filter(Employee.id == d.employee_id).first()
    reportee_name = reportee.full_name if reportee else "Reportee"
    _notify(
        db,
        d.approver_id,
        d.id,
        "kra_submitted",
        f"KRAs submitted for review - {reportee_name}",
        f"Manager has submitted KRAs of {reportee_name} for review.",
    )
    db.commit()
    return {"message": "KRA submitted to approver"}

# ─── Self Rating (Reportee - Feature 3a) ────────────────────────────────────────

@router.post("/diaries/{diary_id}/self-rating")
def save_self_rating(diary_id: str, ratings: List[SelfRatingUpdate], db: Session = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    d = _get_or_404(db, PerformanceDiary, diary_id)
    if d.employee_id != current_user.id: raise HTTPException(403, "Not your diary")
    if d.self_status.value not in ["open"]: raise HTTPException(400, "Self-rating window not open")
    kra_ids = [r.kra_id for r in ratings]
    kras = {
        str(k.id): k for k in db.query(DiaryKRA).filter(DiaryKRA.diary_id == diary_id, DiaryKRA.id.in_(kra_ids)).all()
    } if kra_ids else {}
    for r in ratings:
        if not r.self_comments or not r.self_comments.strip():
            raise HTTPException(400, "Self comments are mandatory")
        kra = kras.get(str(r.kra_id))
        if kra:
            kra.self_rating = r.self_rating
            kra.self_comments = r.self_comments
    db.commit()
    return {"message": "Self ratings saved"}

@router.post("/diaries/{diary_id}/submit-self-rating")
def submit_self_rating(diary_id: str, db: Session = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    d = _get_or_404(db, PerformanceDiary, diary_id)
    if d.employee_id != current_user.id: raise HTTPException(403, "Not your diary")
    if d.self_status.value != "open": raise HTTPException(400, "Cannot submit")
    d.self_status = SelfStatus.submitted
    d.self_submitted_at = datetime.utcnow()
    cycle = d.cycle or db.query(PerformanceCycle).filter(PerformanceCycle.id == d.cycle_id).first()
    mgr_deadline = cycle.step2_mgr_deadline.strftime("%Y-%m-%d") if cycle and cycle.step2_mgr_deadline else "the deadline"
    approval_deadline = cycle.step2_approval_date.strftime("%Y-%m-%d") if cycle and cycle.step2_approval_date else "the deadline"
    reportee_name = current_user.full_name
    _notify_unique(
        db,
        [d.manager_id, d.approver_id],
        d.id,
        "self_rating_submitted",
        f"Self-rating submitted - {reportee_name}",
        (
            f"{reportee_name} has submitted their self-rating.\n"
            f"- Managers: Please complete manager ratings by {mgr_deadline}\n"
            f"- Approver: Please complete review & approval process by {approval_deadline}"
        ),
    )
    db.commit()
    return {"message": "Self-rating submitted"}

# ─── Manager Rating (Feature 3b) ────────────────────────────────────────────────

@router.post("/diaries/{diary_id}/manager-rating")
def save_mgr_rating(diary_id: str, ratings: List[ManagerRatingUpdate], db: Session = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    d = _get_or_404(db, PerformanceDiary, diary_id)
    if d.manager_id != current_user.id: raise HTTPException(403, "Not authorised")
    if d.self_status.value not in ["submitted", "auto_submitted"]: raise HTTPException(400, "Reportee has not submitted yet")
    if d.final_status in [FinalStatus.baselined, FinalStatus.closed]:
        raise HTTPException(400, "Final rating is already locked")
    if d.mgr_status == MgrStatus.submitted and not _can_manager_rerate(db, d):
        raise HTTPException(400, "Manager rating already submitted")
    kra_ids = [r.kra_id for r in ratings]
    kras = {
        str(k.id): k for k in db.query(DiaryKRA).filter(DiaryKRA.diary_id == diary_id, DiaryKRA.id.in_(kra_ids)).all()
    } if kra_ids else {}
    for r in ratings:
        if not r.mgr_comments: raise HTTPException(400, "Manager comments are mandatory")
        kra = kras.get(str(r.kra_id))
        if kra:
            kra.mgr_rating = r.mgr_rating
            kra.mgr_comments = r.mgr_comments
    db.commit()
    return {"message": "Manager ratings saved"}

@router.post("/diaries/{diary_id}/submit-manager-rating")
def submit_mgr_rating(diary_id: str, db: Session = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    d = _get_or_404(db, PerformanceDiary, diary_id)
    if d.manager_id != current_user.id: raise HTTPException(403, "Not authorised")
    if d.final_status in [FinalStatus.baselined, FinalStatus.closed]:
        raise HTTPException(400, "Final rating is already locked")
    if d.mgr_status == MgrStatus.submitted and not _can_manager_rerate(db, d):
        raise HTTPException(400, "Manager rating already submitted")
    kras = db.query(DiaryKRA).filter(DiaryKRA.diary_id == diary_id).all()
    for kra in kras:
        if not kra.mgr_rating or not kra.mgr_comments:
            raise HTTPException(400, "All KRAs must have manager rating and comments")
    d.mgr_status = MgrStatus.submitted
    d.mgr_submitted_at = datetime.utcnow()
    reportee = db.query(Employee).filter(Employee.id == d.employee_id).first()
    reportee_name = reportee.full_name if reportee else "Reportee"
    cycle = d.cycle or db.query(PerformanceCycle).filter(PerformanceCycle.id == d.cycle_id).first()
    approval_deadline = cycle.step2_approval_date.strftime("%Y-%m-%d") if cycle and cycle.step2_approval_date else "the deadline"
    _notify_unique(
        db,
        [d.approver_id],
        d.id,
        "mgr_rating_submitted",
        f"Manager rating submitted - {reportee_name}",
        (
            f"{current_user.full_name} has submitted manager ratings for {reportee_name}.\n"
            f"- Approver: Please complete review & approval process by {approval_deadline}"
        ),
    )
    db.commit()
    return {"message": "Manager rating submitted to approver"}

# ─── Approver Actions ──────────────────────────────────────────────────────────

@router.post("/diaries/{diary_id}/approve-kra")
def approve_kra(diary_id: str, data: ApprovalActionRequest, db: Session = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    d = _get_or_404(db, PerformanceDiary, diary_id)
    if d.approver_id != current_user.id and current_user.role != RoleEnum.admin:
        raise HTTPException(403, "Not the approver")
    if d.kra_status == DiaryKRAStatus.baselined:
        raise HTTPException(400, "KRA is already baselined")
    if d.final_status in [FinalStatus.baselined, FinalStatus.closed]:
        raise HTTPException(400, "Performance diary is already finalised")
    if data.action == "approve":
        if d.kra_status != DiaryKRAStatus.submitted:
            raise HTTPException(400, "Manager must resubmit KRA before approval")
        d.kra_status = DiaryKRAStatus.baselined
        d.kra_baselined_at = datetime.utcnow()
        db.add(ApprovalAction(
            id=str(uuid.uuid4()),
            diary_id=d.id,
            actor_id=current_user.id,
            action_type=ActionType.approve,
            stage=ActionStage.kra,
            comment=data.comment,
        ))
        reportee = db.query(Employee).filter(Employee.id == d.employee_id).first()
        reportee_name = reportee.full_name if reportee else "Reportee"
        baselined_title = f"KRA Baselined - {reportee_name}"
        baselined_body = f"KRA for {reportee_name} has been approved & baselined."
        _notify(db, d.employee_id, d.id, "kra_baselined", baselined_title, baselined_body)
        _notify(db, d.manager_id, d.id, "kra_baselined", baselined_title, baselined_body)
    elif data.action == "send_back":
        if not data.comment or not data.comment.strip():
            raise HTTPException(400, "Comment is required for send-back")
        if d.kra_sendback_count >= 2: raise HTTPException(400, "Maximum send-backs (2) reached")
        if d.kra_status != DiaryKRAStatus.submitted:
            raise HTTPException(400, "KRA must be submitted before it can be sent back")
        d.kra_status = DiaryKRAStatus.sent_back
        d.kra_sendback_count += 1
        db.add(ApprovalAction(
            id=str(uuid.uuid4()),
            diary_id=d.id,
            actor_id=current_user.id,
            action_type=ActionType.send_back,
            stage=ActionStage.kra,
            comment=data.comment,
            sendback_seq=d.kra_sendback_count,
        ))
        reportee = db.query(Employee).filter(Employee.id == d.employee_id).first()
        reportee_name = reportee.full_name if reportee else "Reportee"
        cycle = d.cycle or db.query(PerformanceCycle).filter(PerformanceCycle.id == d.cycle_id).first()
        cycle_label = _cycle_label(cycle.cycle_name, cycle.period, bold=True) if cycle else "**the performance cycle**"
        approver_name = current_user.full_name or "Approver"
        _notify_async_email(
            db,
            d.manager_id,
            d.id,
            "kra_sent_back",
            f"KRA Sent Back - {reportee_name}",
            (
                f"{approver_name} has sent back the KRA for the performance cycle {cycle_label}.\n\n"
                f"Approver feedback: {data.comment}"
            ),
        )
    else:
        raise HTTPException(400, "Unsupported approval action")
    db.commit()
    return {"message": f"KRA {data.action}d"}

@router.post("/diaries/{diary_id}/approve-rating")
def approve_rating(diary_id: str, data: ApprovalActionRequest, db: Session = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    d = _get_or_404(db, PerformanceDiary, diary_id)
    if d.approver_id != current_user.id: raise HTTPException(403, "Not the approver")
    if d.final_status in [FinalStatus.baselined, FinalStatus.closed]:
        raise HTTPException(400, "Performance rating is already finalised")
    if data.action == "approve":
        if d.final_review_open:
            raise HTTPException(400, "Final performance review is already open with the manager")
        if d.mgr_status != MgrStatus.submitted:
            raise HTTPException(400, "Manager must resubmit rating before approval")
        employee = db.query(Employee).filter(Employee.id == d.employee_id).first()
        d.final_review_open = True
        d.final_status = FinalStatus.pending
        d.overall_performance_rating = None
        d.overall_performance_comments = None
        db.add(ApprovalAction(
            id=str(uuid.uuid4()),
            diary_id=d.id,
            actor_id=current_user.id,
            action_type=ActionType.approve,
            stage=ActionStage.mgr_rating,
            comment=data.comment,
        ))
        reportee_name = employee.full_name if employee else "Reportee"
        cycle = d.cycle or db.query(PerformanceCycle).filter(PerformanceCycle.id == d.cycle_id).first()
        cycle_label = _cycle_label(cycle.cycle_name, cycle.period, bold=True) if cycle else "the performance cycle"
        cycle_close_date = cycle.step2_approval_date.strftime("%Y-%m-%d") if cycle and cycle.step2_approval_date else "the cycle close date"
        _notify(
            db,
            d.manager_id,
            d.id,
            "final_review_open",
            f"Rating Approved - {reportee_name}",
            (
                f"The performance rating for cycle {cycle_label} has been reviewed and finalised by the approver.\n\n"
                f"The manager must complete the Final Performance Review by providing the overall rating and feedback before the cycle closes on {cycle_close_date}."
            ),
        )
    elif data.action == "send_back":
        if not data.comment or not data.comment.strip():
            raise HTTPException(400, "Comment is required for send-back")
        if d.mgr_sendback_count >= 2: raise HTTPException(400, "Maximum send-backs (2) reached")
        if d.mgr_status != MgrStatus.submitted:
            raise HTTPException(400, "Rating must be submitted before it can be sent back")
        d.mgr_status = MgrStatus.sent_back
        d.final_review_open = False
        d.overall_performance_rating = None
        d.overall_performance_comments = None
        d.mgr_sendback_count += 1
        db.add(ApprovalAction(
            id=str(uuid.uuid4()),
            diary_id=d.id,
            actor_id=current_user.id,
            action_type=ActionType.send_back,
            stage=ActionStage.mgr_rating,
            comment=data.comment,
            sendback_seq=d.mgr_sendback_count,
        ))
        reportee = db.query(Employee).filter(Employee.id == d.employee_id).first()
        reportee_name = reportee.full_name if reportee else "Reportee"
        cycle = d.cycle or db.query(PerformanceCycle).filter(PerformanceCycle.id == d.cycle_id).first()
        mgr_deadline = cycle.step2_mgr_deadline.strftime("%Y-%m-%d") if cycle and cycle.step2_mgr_deadline else "the deadline"
        _notify(
            db,
            d.manager_id,
            d.id,
            "rating_sent_back",
            f"Rating Sent Back - {reportee_name}",
            f"Approver feedback: {data.comment}\n- Managers: Please complete manager ratings by {mgr_deadline}",
        )
    else:
        raise HTTPException(400, "Unsupported approval action")
    db.commit()
    return {"message": f"Rating {data.action}d"}

@router.post("/diaries/{diary_id}/submit-final-review")
def submit_final_review(diary_id: str, data: FinalPerformanceReviewSubmit, db: Session = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    d = _get_or_404(db, PerformanceDiary, diary_id)
    if d.manager_id != current_user.id: raise HTTPException(403, "Not authorised")
    if not d.final_review_open: raise HTTPException(400, "Final performance review is not open")
    if d.mgr_status != MgrStatus.submitted: raise HTTPException(400, "Manager rating must be submitted first")
    if data.overall_performance_rating not in [1, 2, 3, 4, 5]:
        raise HTTPException(400, "Overall performance rating must be between 1 and 5")
    if not data.overall_performance_comments or not data.overall_performance_comments.strip():
        raise HTTPException(400, "Overall performance comment is required")
    employee = db.query(Employee).filter(Employee.id == d.employee_id).first()
    d.overall_performance_rating = data.overall_performance_rating
    d.overall_performance_comments = data.overall_performance_comments.strip()
    d.final_review_open = False
    d.final_status = FinalStatus.baselined
    d.final_baselined_at = datetime.utcnow()
    reportee_name = employee.full_name if employee else "Reportee"
    cycle = d.cycle or db.query(PerformanceCycle).filter(PerformanceCycle.id == d.cycle_id).first()
    cycle_label = _cycle_label(cycle.cycle_name, cycle.period, bold=True) if cycle else "the performance cycle"
    _notify_unique(
        db,
        [d.employee_id, d.approver_id],
        d.id,
        "rating_baselined",
        f"Final Performance Review Submitted - {reportee_name}",
        (
            f"The performance feedback for cycle {cycle_label} has been submitted by {current_user.full_name}.\n\n"
            f"You may now log in to the EDGE portal to review the final rating and feedback."
        ),
    )
    db.commit()
    return {"message": "Final performance review submitted"}

# ─── Grievances ────────────────────────────────────────────────────────────────

@router.get("/grievances/my")
def my_grievances(page: Optional[int] = None, page_size: int = 25, db: Session = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    q = db.query(Grievance).filter(Grievance.raised_by == current_user.id).options(joinedload(Grievance.raised_by_user)).order_by(Grievance.raised_at.desc())
    if page:
        page = max(page, 1)
        page_size = min(max(page_size, 1), 100)
        total = q.count()
        gs = q.offset((page - 1) * page_size).limit(page_size).all()
        return {"items": [_grievance_dict(g) for g in gs], "total": total, "page": page, "page_size": page_size, "total_pages": (total + page_size - 1) // page_size}
    return [_grievance_dict(g) for g in q.all()]

@router.get("/grievances/team")
def team_grievances(page: Optional[int] = None, page_size: int = 25, db: Session = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    diary_ids = [d.id for d in db.query(PerformanceDiary).filter(
        or_(PerformanceDiary.manager_id == current_user.id, PerformanceDiary.approver_id == current_user.id),
        PerformanceDiary.employee_id != current_user.id
    ).all()]
    q = db.query(Grievance).filter(Grievance.diary_id.in_(diary_ids)).options(joinedload(Grievance.raised_by_user)).order_by(Grievance.raised_at.desc())
    if page:
        page = max(page, 1)
        page_size = min(max(page_size, 1), 100)
        total = q.count()
        gs = q.offset((page - 1) * page_size).limit(page_size).all()
        return {"items": [_grievance_dict(g) for g in gs], "total": total, "page": page, "page_size": page_size, "total_pages": (total + page_size - 1) // page_size}
    return [_grievance_dict(g) for g in q.all()]

@router.get("/grievances/approver")
def approver_grievances(page: Optional[int] = None, page_size: int = 25, db: Session = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    diary_ids = [d.id for d in db.query(PerformanceDiary).filter(PerformanceDiary.approver_id == current_user.id).all()]
    q = db.query(Grievance).filter(Grievance.diary_id.in_(diary_ids)).options(joinedload(Grievance.raised_by_user)).order_by(Grievance.raised_at.desc())
    if page:
        page = max(page, 1)
        page_size = min(max(page_size, 1), 100)
        total = q.count()
        gs = q.offset((page - 1) * page_size).limit(page_size).all()
        return {"items": [_grievance_dict(g) for g in gs], "total": total, "page": page, "page_size": page_size, "total_pages": (total + page_size - 1) // page_size}
    return [_grievance_dict(g) for g in q.all()]

@router.post("/grievances")
def raise_grievance(data: GrievanceCreate, db: Session = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    diary = None
    if data.diary_kra_id:
        kra = db.query(DiaryKRA).filter(DiaryKRA.id == data.diary_kra_id).first()
        diary = db.query(PerformanceDiary).filter(PerformanceDiary.id == kra.diary_id).first() if kra else None
    if not diary:
        diary = db.query(PerformanceDiary).filter(PerformanceDiary.employee_id == current_user.id).order_by(PerformanceDiary.created_at.desc()).first()
    if not diary: raise HTTPException(404, "No diary found")
    g = Grievance(id=str(uuid.uuid4()), diary_id=diary.id, 
                  diary_kra_id=data.diary_kra_id if data.diary_kra_id else None,
                  raised_by=current_user.id, grievance_type=data.grievance_type,
                  description=data.description, status=GrievanceStatus.l1_review,
                  current_level=1, sla_due_at=datetime.utcnow() + timedelta(days=7))
    db.add(g)
    _notify(db, diary.manager_id, diary.id, "grievance_raised", "Grievance Raised", f"Employee raised a grievance: {data.description[:100]}")
    _notify(db, diary.approver_id, diary.id, "grievance_raised", "Grievance Raised", f"Employee raised a grievance: {data.description[:100]}")
    db.commit()
    return _grievance_dict(g)

@router.post("/grievances/{grievance_id}/respond")
def respond_grievance(grievance_id: str, data: GrievanceRespond, db: Session = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    g = _get_or_404(db, Grievance, grievance_id)
    now = datetime.utcnow()
    diary = db.query(PerformanceDiary).filter(PerformanceDiary.id == g.diary_id).first()
    if g.current_level == 1:
        g.l1_response = data.response; g.l1_responded_at = now
        if data.resolve:
            g.status = GrievanceStatus.resolved; g.resolved_at = now
        else:
            g.current_level = 2; g.status = GrievanceStatus.l2_review
            g.sla_due_at = now + timedelta(days=7)
    elif g.current_level == 2:
        g.l2_response = data.response; g.l2_responded_at = now
        if data.resolve:
            g.status = GrievanceStatus.resolved; g.resolved_at = now
        else:
            g.current_level = 3; g.status = GrievanceStatus.l3_review
            g.sla_due_at = now + timedelta(days=7)
    elif g.current_level == 3:
        g.l3_response = data.response; g.l3_responded_at = now
        g.status = GrievanceStatus.resolved; g.resolved_at = now
    if diary:
        _notify(db, diary.employee_id, diary.id, "grievance_response", "Grievance Update", f"Your grievance has received a response at Level {g.current_level}")
    db.commit()
    return _grievance_dict(g)

# ─── Notifications ─────────────────────────────────────────────────────────────

@router.get("/notifications")
def my_notifications(page: Optional[int] = None, page_size: int = 25, db: Session = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    q = db.query(Notification).filter(Notification.recipient_id == current_user.id).order_by(Notification.created_at.desc())
    total = None
    if page:
        page = max(page, 1)
        page_size = min(max(page_size, 1), 100)
        total = q.count()
        notifs = q.offset((page - 1) * page_size).limit(page_size).all()
    else:
        notifs = q.limit(50).all()
    diary_ids = {n.diary_id for n in notifs if n.diary_id}
    diaries = {
        d.id: d for d in db.query(PerformanceDiary).filter(PerformanceDiary.id.in_(diary_ids)).all()
    } if diary_ids else {}
    employee_ids = {d.employee_id for d in diaries.values()}
    employees = {
        e.id: e for e in db.query(Employee).filter(Employee.id.in_(employee_ids)).all()
    } if employee_ids else {}
    result = []
    for n in notifs:
        emp_name = "System"
        if n.diary_id:
            diary = diaries.get(n.diary_id)
            if diary:
                emp = employees.get(diary.employee_id)
                if emp:
                    emp_name = emp.full_name
        result.append({
            "id": n.id, "event_type": n.event_type, "title": n.title, "body": n.body,
            "is_read": n.is_read, "created_at": n.created_at.isoformat(),
            "diary_id": n.diary_id,
            "employee_name": emp_name
        })
    if page:
        return {
            "items": result,
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": (total + page_size - 1) // page_size,
        }
    return result

@router.get("/notifications/unread-count")
def notification_unread_count(db: Session = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    unread = db.query(func.count(Notification.id)).filter(
        Notification.recipient_id == current_user.id,
        Notification.is_read == False,
    ).scalar()
    return {"unread": unread or 0}

@router.post("/notifications/{notif_id}/read")
def mark_read(notif_id: str, db: Session = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    n = db.query(Notification).filter(Notification.id == notif_id, Notification.recipient_id == current_user.id).first()
    if n: n.is_read = True; db.commit()
    return {"message": "Marked as read"}

@router.post("/notifications/read-all")
def mark_all_read(db: Session = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    db.query(Notification).filter(Notification.recipient_id == current_user.id, Notification.is_read == False).update({"is_read": True})
    db.commit()
    return {"message": "All marked as read"}

# ─── Admin ─────────────────────────────────────────────────────────────────────

@router.get("/admin/dashboard")
def admin_dashboard(db: Session = Depends(get_db), _=Depends(require_role("admin"))):
    active_cycle = db.query(PerformanceCycle).filter(PerformanceCycle.status.in_(["step1","step2"])).first()
    active_cycle_stats = None
    if active_cycle:
        cycle_diaries = (
            db.query(PerformanceDiary)
            .join(Employee, PerformanceDiary.employee_id == Employee.id)
            .filter(
                PerformanceDiary.cycle_id == active_cycle.id,
                Employee.is_active == True,
            )
            .all()
        )
        active_cycle_stats = {
            "cycle_id": active_cycle.id,
            "cycle_name": active_cycle.cycle_name,
            "total_diaries": len(cycle_diaries),
            "kra_pending": sum(1 for d in cycle_diaries if d.kra_status != DiaryKRAStatus.baselined),
            "kra_submitted": sum(1 for d in cycle_diaries if d.kra_status == DiaryKRAStatus.submitted),
            "self_rating_done": sum(1 for d in cycle_diaries if d.self_status in [SelfStatus.submitted, SelfStatus.auto_submitted]),
            "manager_rating_done": sum(1 for d in cycle_diaries if d.mgr_status == MgrStatus.submitted),
            "approver_rating_done": sum(1 for d in cycle_diaries if d.final_review_open or d.final_status in [FinalStatus.baselined, FinalStatus.closed]),
            "final_review_done": sum(1 for d in cycle_diaries if d.final_status in [FinalStatus.baselined, FinalStatus.closed]),
            "final_baselined": sum(1 for d in cycle_diaries if d.final_status == FinalStatus.baselined),
            "final_closed": sum(1 for d in cycle_diaries if d.final_status == FinalStatus.closed),
            "kra_baselined": sum(1 for d in cycle_diaries if d.kra_status == DiaryKRAStatus.baselined),
            "ratings_pending": sum(1 for d in cycle_diaries if d.mgr_status == MgrStatus.submitted and not d.final_review_open and d.final_status not in [FinalStatus.baselined, FinalStatus.closed]),
        }
    return {
        "total_employees": db.query(Employee).filter(Employee.is_active == True).count(),
        "open_cycles": db.query(PerformanceCycle).filter(PerformanceCycle.status.in_(["draft", "step1", "step2"])).count(),
        "closed_cycles": db.query(PerformanceCycle).filter(PerformanceCycle.status == CycleStatus.closed).count(),
        "active_cycle": _cycle_dict(active_cycle),
        "active_cycle_stats": active_cycle_stats,
        "total_diaries": active_cycle_stats["total_diaries"] if active_cycle_stats else 0,
        "open_grievances": db.query(Grievance).filter(Grievance.status.in_(["open","l1_review","l2_review","l3_review"])).count(),
    }

# ─── Helpers ───────────────────────────────────────────────────────────────────

@router.get("/admin/employee-status/export")
def export_employee_status(
    cycle_id: Optional[str] = None,
    status_filter: Optional[str] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    _=Depends(require_role("admin")),
):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "Excel export dependency is missing. Install openpyxl.")

    q = (
        db.query(PerformanceDiary)
        .join(Employee, PerformanceDiary.employee_id == Employee.id)
        .options(joinedload(PerformanceDiary.cycle))
        .filter(Employee.is_active == True)
    )
    if cycle_id:
        q = q.filter(PerformanceDiary.cycle_id == cycle_id)
    diaries = q.all()

    people_ids = {
        person_id
        for d in diaries
        for person_id in (d.employee_id, d.manager_id, d.approver_id)
        if person_id
    }
    people = {
        e.id: e for e in db.query(Employee).options(joinedload(Employee.band)).filter(Employee.id.in_(people_ids)).all()
    } if people_ids else {}

    rows = []
    for d in diaries:
        emp = people.get(d.employee_id)
        if not emp:
            continue
        manager = people.get(d.manager_id)
        approver = people.get(d.approver_id)
        if search and not any(
            _matches_employee_search(person, search)
            for person in [emp, manager, approver]
        ):
            continue
        if status_filter and not _matches_admin_status_filter(d, status_filter):
            continue
        rows.append([
            emp.employee_code or "",
            emp.full_name or "",
            emp.email or "",
            emp.band_code or (emp.band.band_code if emp.band else "") or "",
            manager.full_name if manager else "",
            manager.email if manager else "",
            approver.full_name if approver else "",
            approver.email if approver else "",
            _employee_status_final_label(d),
            _employee_status_current_step_label(d),
        ])

    rows.sort(key=lambda row: (row[1] or "").lower())
    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Status"
    headers = [
        "Employee_code",
        "Employee",
        "Employee_mail",
        "Band",
        "Manager",
        "Manager_mail",
        "Approver",
        "Approver_mail",
        "Final Status",
        "Current Status",
    ]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="EAF0FF")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="0F3F9F")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(header)
        for cell in ws[get_column_letter(col_idx)]:
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 42)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"employee_status_{timestamp}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

def _get_or_404(db, model, id):
    obj = db.query(model).filter(model.id == id).first()
    if not obj: raise HTTPException(404, f"{model.__name__} not found")
    return obj

def _emp_dict(e):
    if not e: return None
    band_code = e.band_code or (e.band.band_code if e.band else None)
    band_name = e.band_name or (e.band.band_name if e.band else None)
    return {"id": e.id, "employee_code": e.employee_code, "full_name": e.full_name, "email": e.email,
            "role": e.role.value, "band_id": e.band_id, "band_code": band_code, "band_name": band_name, "manager_id": e.manager_id, "approver_id": e.approver_id}

def _cycle_dict(c):
    if not c: return None
    raw_period = c.period.value if hasattr(c.period, 'value') else c.period
    return {"id": c.id, "cycle_name": c.cycle_name, "financial_year": c.financial_year, "period": DB_TO_PUBLIC_CYCLE_PERIOD.get(raw_period, raw_period),
            "status": c.status.value if hasattr(c.status, 'value') else c.status,
            "created_at": c.created_at.isoformat() if c.created_at else None,
            "step1_open_date": c.step1_open_date.strftime("%Y-%m-%d"), "step1_kra_deadline": c.step1_kra_deadline.strftime("%Y-%m-%d"),
            "step1_approval_date": c.step1_approval_date.strftime("%Y-%m-%d"),
            "step2_open_date": c.step2_open_date.strftime("%Y-%m-%d") if c.step2_open_date else None,
            "step2_self_deadline": c.step2_self_deadline.strftime("%Y-%m-%d") if c.step2_self_deadline else None,
            "step2_mgr_deadline": c.step2_mgr_deadline.strftime("%Y-%m-%d") if c.step2_mgr_deadline else None,
            "step2_approval_date": c.step2_approval_date.strftime("%Y-%m-%d") if c.step2_approval_date else None}

def _enum_value(value):
    return value.value if hasattr(value, "value") else value

def _employee_status_rating_started(d):
    cycle_status = _enum_value(d.cycle.status) if d.cycle else None
    return (
        cycle_status == "step2"
        or _enum_value(d.self_status) != "not_open"
        or _enum_value(d.mgr_status) != "pending"
        or bool(d.final_review_open)
        or _enum_value(d.final_status) in ["baselined", "closed"]
    )

def _employee_status_final_label(d):
    final_status = _enum_value(d.final_status)
    if final_status == "closed":
        return "Closed"
    if final_status == "baselined":
        return "Baselined"
    if not _employee_status_rating_started(d):
        return "Not started"
    if d.final_review_open:
        return "Final review"
    return "Pending"

def _employee_status_current_step_label(d):
    final_status = _enum_value(d.final_status)
    kra_status = _enum_value(d.kra_status)
    mgr_status = _enum_value(d.mgr_status)
    self_status = _enum_value(d.self_status)

    if final_status in ["closed", "baselined"]:
        return "Completed"
    if kra_status == "submitted":
        return "Step 1 - Approver review"
    if kra_status == "sent_back":
        return "Step 1 - KRA rework"
    if kra_status != "baselined":
        return "Step 1 - KRA allocation"
    if mgr_status == "sent_back":
        return "Step 2 - Manager rework"
    if mgr_status == "submitted":
        if d.final_review_open:
            return "Step 2 - Final manager review"
        return "Step 2 - Approver review"
    if self_status in ["submitted", "auto_submitted"]:
        return "Step 2 - Manager rating"
    if self_status == "open":
        return "Step 2 - Self rating"
    return "Step 1 - KRA allocated"

def _matches_admin_status_filter(d, status_filter):
    s = (status_filter or "").lower()
    if not s:
        return True
    is_completed = _enum_value(d.final_status) in ["baselined", "closed"]
    is_rating_started = _employee_status_rating_started(d)
    is_step2 = not is_completed and _enum_value(d.kra_status) == "baselined" and is_rating_started
    is_step1 = not is_completed and not is_step2

    if s == "step1":
        return is_step1
    if s == "step2":
        return is_step2
    if s == "completed":
        return is_completed
    if s == "pending":
        return _enum_value(d.final_status) == "pending"
    return True

def _diary_dict(d, db: Session | None = None):
    if not d: return None
    has_open_grievance = False
    active_sendback = {"kra": None, "rating": None}
    kra_rework_history = []
    if db:
        has_open_grievance = db.query(Grievance).filter(
            Grievance.diary_id == d.id,
            Grievance.status.notin_([GrievanceStatus.resolved, GrievanceStatus.closed])
        ).first() is not None
        kra_actions = db.query(ApprovalAction).filter(
            ApprovalAction.diary_id == d.id,
            ApprovalAction.stage == ActionStage.kra,
            ApprovalAction.action_type == ActionType.send_back,
        ).order_by(ApprovalAction.actioned_at.asc()).all()
        actor_ids = {action.actor_id for action in kra_actions if action.actor_id}
        actors = {
            employee.id: employee.full_name for employee in db.query(Employee).filter(Employee.id.in_(actor_ids)).all()
        } if actor_ids else {}
        kra_rework_history = [
            {
                "action_type": action.action_type.value,
                "actor_id": action.actor_id,
                "actor_name": actors.get(action.actor_id, "Approver"),
                "comment": action.comment,
                "sendback_seq": action.sendback_seq,
                "actioned_at": action.actioned_at.isoformat() if action.actioned_at else None,
            }
            for action in kra_actions
        ]
        if d.kra_status == DiaryKRAStatus.sent_back:
            action = kra_actions[-1] if kra_actions else None
            if action:
                active_sendback["kra"] = {
                    "comment": action.comment,
                    "actioned_at": action.actioned_at.isoformat() if action.actioned_at else None,
                }
        if d.mgr_status == MgrStatus.sent_back:
            action = db.query(ApprovalAction).filter(
                ApprovalAction.diary_id == d.id,
                ApprovalAction.stage == ActionStage.mgr_rating,
                ApprovalAction.action_type == ActionType.send_back,
            ).order_by(ApprovalAction.actioned_at.desc()).first()
            if action:
                active_sendback["rating"] = {
                    "comment": action.comment,
                    "actioned_at": action.actioned_at.isoformat() if action.actioned_at else None,
                }
    return {"id": d.id, "cycle_id": d.cycle_id, "employee_id": d.employee_id, "manager_id": d.manager_id, "approver_id": d.approver_id,
            "kra_status": d.kra_status.value, "kra_sendback_count": d.kra_sendback_count, "kra_baselined_at": d.kra_baselined_at.isoformat() if d.kra_baselined_at else None,
            "self_status": d.self_status.value, "self_submitted_at": d.self_submitted_at.isoformat() if d.self_submitted_at else None,
            "mgr_status": d.mgr_status.value, "mgr_sendback_count": d.mgr_sendback_count, "mgr_submitted_at": d.mgr_submitted_at.isoformat() if d.mgr_submitted_at else None,
            "final_status": d.final_status.value, "final_review_open": bool(d.final_review_open),
            "overall_performance_rating": d.overall_performance_rating,
            "overall_performance_comments": d.overall_performance_comments,
            "final_baselined_at": d.final_baselined_at.isoformat() if d.final_baselined_at else None,
            "created_at": d.created_at.isoformat() if d.created_at else None,
            "updated_at": d.updated_at.isoformat() if d.updated_at else None,
            "has_open_grievance": has_open_grievance,
            "active_sendback": active_sendback,
            "kra_rework_history": kra_rework_history,
            "cycle": _cycle_dict(d.cycle) if d.cycle else None}

def _grievance_dict(g):
    return {
        "id": g.id, "diary_id": g.diary_id, "diary_kra_id": g.diary_kra_id,
        "raised_by": {"id": str(g.raised_by_user.id), "name": g.raised_by_user.full_name} if g.raised_by_user else None,
        "employee_name": g.raised_by_user.full_name if g.raised_by_user else "Unknown",
        "grievance_type": g.grievance_type, "description": g.description, "status": g.status.value,
        "current_level": g.current_level, "l1_response": g.l1_response, "l2_response": g.l2_response,
        "l3_response": g.l3_response, "sla_due_at": g.sla_due_at.isoformat() if g.sla_due_at else None,
        "raised_at": g.raised_at.isoformat(), "resolved_at": g.resolved_at.isoformat() if g.resolved_at else None
    }

def _open_grievance_diary_ids(db, diary_ids):
    if not diary_ids:
        return set()
    rows = db.query(Grievance.diary_id).filter(
        Grievance.diary_id.in_(diary_ids),
        Grievance.status.notin_([GrievanceStatus.resolved, GrievanceStatus.closed])
    ).distinct().all()
    return {row[0] for row in rows}

def _active_employee_recipient_ids(db, recipient_ids):
    recipient_ids = {recipient_id for recipient_id in recipient_ids if recipient_id}
    if not recipient_ids:
        return set()
    rows = db.query(Employee.id).filter(
        Employee.id.in_(recipient_ids),
        Employee.is_active == True,
    ).all()
    return {row[0] for row in rows}

def _notify(db, recipient_id, diary_id, event_type, title, body):
    if recipient_id not in _active_employee_recipient_ids(db, [recipient_id]):
        return
    n = Notification(id=str(uuid.uuid4()), recipient_id=recipient_id, diary_id=diary_id,
                     event_type=event_type, title=title, body=body, channel=NotifChannel.both)
    db.add(n)
    _send_emails_for_notifications(db, [n])

def _notify_async_email(db, recipient_id, diary_id, event_type, title, body):
    if recipient_id not in _active_employee_recipient_ids(db, [recipient_id]):
        return
    n = Notification(id=str(uuid.uuid4()), recipient_id=recipient_id, diary_id=diary_id,
                     event_type=event_type, title=title, body=body, channel=NotifChannel.both)
    db.add(n)
    _send_emails_for_notifications_async(db, [n])

def _notify_unique(db, recipient_ids, diary_id, event_type, title, body):
    seen = set()
    for recipient_id in recipient_ids:
        if recipient_id and recipient_id not in seen:
            seen.add(recipient_id)
            _notify(db, recipient_id, diary_id, event_type, title, body)

def _send_emails_for_notifications(db, notifications):
    payloads = _email_payloads_for_notifications(db, notifications)
    _send_email_payloads(payloads)

def _send_emails_for_notifications_async(db, notifications):
    payloads = _email_payloads_for_notifications(db, notifications)
    if payloads:
        email_executor.submit(_send_email_payloads, payloads)

def _email_payloads_for_notifications(db, notifications):
    if not notifications:
        return []
    recipient_ids = {n.recipient_id for n in notifications if n.recipient_id}
    recipients = {
        e.id: e.email for e in db.query(Employee.id, Employee.email).filter(
            Employee.id.in_(recipient_ids),
            Employee.is_active == True,
        ).all()
    } if recipient_ids else {}
    payloads = []
    for n in notifications:
        if n.event_type not in EMAIL_ENABLED_EVENT_TYPES:
            continue
        email = recipients.get(n.recipient_id)
        if email:
            payloads.append({
                "notification_id": n.id,
                "to_email": email,
                "title": n.title,
                "body": n.body,
            })
    return payloads

def _send_email_payloads(payloads):
    try:
        sent_ids = send_notification_email_batch(payloads)
    except Exception as exc:
        print(f"Email notification batch failed: {exc}")
        sent_ids = []
    if not sent_ids:
        return
    db = SessionLocal()
    try:
        notifications = db.query(Notification).filter(Notification.id.in_(sent_ids)).all()
        for n in notifications:
            mark_email_sent(n)
        db.commit()
    except Exception as exc:
        db.rollback()
        print(f"Email sent status update failed: {exc}")
    finally:
        db.close()

def _send_emails_for_notifications_legacy(db, notifications):
    for n in notifications:
        try:
            recipient = db.query(Employee).filter(
                Employee.id == n.recipient_id,
                Employee.is_active == True,
            ).first()
            if recipient and send_notification_email(recipient.email, n.title, n.body):
                mark_email_sent(n)
        except Exception as exc:
            print(f"Email notification failed for {n.recipient_id}: {exc}")

def _has_open_grievance(db, diary_id):
    return db.query(Grievance).filter(
        Grievance.diary_id == diary_id,
        Grievance.status.notin_([GrievanceStatus.resolved, GrievanceStatus.closed])
    ).first() is not None

def _can_manager_rerate(db, diary):
    return (
        diary.mgr_status == MgrStatus.sent_back or
        diary.kra_status == DiaryKRAStatus.sent_back or
        _has_open_grievance(db, diary.id)
    )

def _create_diaries_for_cycle(cycle_id, db):
    employees = db.query(Employee).filter(
        Employee.role.in_([RoleEnum.reportee, RoleEnum.manager, RoleEnum.approver]),
        Employee.is_active == True
    ).all()
    for emp in employees:
        mid = emp.manager_id or emp.approver_id
        aid = emp.approver_id or mid
        if mid and aid:
            existing = db.query(PerformanceDiary).filter(
                PerformanceDiary.cycle_id == cycle_id,
                PerformanceDiary.employee_id == emp.id
            ).first()
            if existing:
                continue
            d = PerformanceDiary(id=str(uuid.uuid4()), cycle_id=cycle_id, employee_id=emp.id,
                                 manager_id=mid, approver_id=aid)
            db.add(d)
    db.commit()
